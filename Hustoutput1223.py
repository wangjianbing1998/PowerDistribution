#coding=utf-8
#加载模块
#from EtoX import writeInfoToXml
import sys
import re
# import xlwt
import xlrd
import xlutils
from xlutils.copy import copy
import os
import win32com.client as win32
from win32com.client import Dispatch
import xlwings
import Path
import time
import openpyxl

# from pyExcelerator import *
#from pandas.core.frame import DataFrame
#import pandas as pd


#该程序用来读取下层电力盈亏以及计算联络线功率、互济后的上下层功率
#s怎么传进去？


class exchange():
	def __init__(self,s,wpl,wex):
		self.s=s
		self.wpl=wpl
		super().__init__()
		def search(path=".", name=""):  #search函数找到想要的几个方案的XML文档添加到result中
			for item in os.listdir(path):
				item_path = os.path.join(path, item)
				if os.path.isdir(item_path):
					search(item_path, name)
				elif os.path.isfile(item_path):
					if name in item:
						global result
						result.append(item_path)
		def link(num):
			for i in range(1,Path.system.shape[0]):
				if int(num) == int(Path.system[i, 3]):
					m = int(Path.system[i, 0])
					return (m)
		def rank_Hj():
			t=[]
			r=[]
			for i in range(1,Path.system.shape[0]):
				if int(Path.system[i,3])>0:
					t.append(int(Path.system[i, 3]))
			r.append(min(t))
			r.append(max(t))
			return (r)

		global result
		result = []
		#TODO

		# filepath='C:\\Users\zhang\Desktop\软件\调试用\Hust数据\\'
		filepath=Path.filepath + 'Hust数据\\'
		filename='南疆电网.xlsx'
		print(filepath)
		# ysb = xlrd.open_workbook(Path.filepath + Path.filename[0:-5] +str(self.s)+ '.xlsx')  # 读入南疆电网第s场景下的结果
		# ysb_sheet5 = ysb.sheet_by_name('系统表')  #取“系统表”这一sheet
		# ysb_sheet5_nrows = ysb_sheet5.nrows #行数存入ysb_sheet6_nrows
		# jzb1 = []   # 分区名称集合
		# for m in range(1, ysb_sheet5_nrows):
		# 	if 100 <= float(ysb_sheet5.cell(m, 2).value) <= 101:
		# 		jzb1.append(ysb_sheet5.cell(m, 1).value)#存分区名称
		# fenqu = len(jzb1) - 1  # 确定分区数目（减去的是节点0：系统）

		# if float(Path.FangAndID) ==0:
		# 	down = 0  # 日类型分区
		# 	up = 0
		# elif float(Path.FangAndID) ==1:
		# 	down = fenqu + 1   #典型周第一天序号
		# 	up = fenqu + 7     #典型周第七天序号
        #
		# # print(down)
		# if float(Path.FangAndID) ==0:
		# 	beizhu='最大负荷日'
		# elif float(Path.FangAndID) ==1:
		# 	beizhu='典型周'

		# book1 = xlrd.open_workbook(Path.filepath +  '云南系统' + Path.shuipingnian+'年' +"全部方案"+'指标统计表' + '.xls', 'w+')  # 以读的方式打开
		# book2 = copy(book1) #复制这个表才可以对已有的表写入
		# book1=Workbook(Path.filepath + Path.filename[0:-5] +str(self.s)+ '_GEN.xlsx')
		print(self.s)
		search(path=filepath, name=filename[0:-5] + str(self.s) +'_GEN.xml')
		print(result)
		for ii in range(len(result)):
			filewz1 = result[ii]
			f = open(filewz1, "r", encoding='UTF-8')  # 以读的方式打开XML文档
			content = f.read()
			f.close()

		book = openpyxl.Workbook()  # 创建一个工作表
		ws = book.active  # ws操作sheet页
		sheet1 = book.create_sheet('电力盈亏', 0)

		# openpyxl.load_workbook
		# book1=self.book
		# sheet1 = book.add_sheet('电力盈亏', cell_overwrite_ok=True)
		# tall_style = openpyxl.easyxf("font:height 400")
		title = ["场景ID", "月份", "H1", "H2", "H3", "H4", "H5", "H6","H7", "H8", "H9", "H10", "H11", "H12", "H13",
				 "H14", "H15", "H16", "H17", "H18", "H19", "H20", "H21", "H22", "H23", "H24","系统ID","备用盈亏","日类型"]
		for i in title:
			sheet1.cell(1, title.index(i)+1, i)

		channelList = re.findall("<HST.*?</HST>", content, re.S)
			# 新建一些空集来存放读取的XML文档的列内容
		list0 = []
		list1 = []
		list2 = []
		list3 = []
		list4 = []
		list5 = []
		list6 = []
		list7 = []
		list8 = []
		list9 = []
		list10 = []
		list11 = []
		list12 = []
		list13 = []
		list14 = []
		list15 = []
		list16 = []
		list17 = []
		list18 = []
		list19 = []
		list20 = []
		list21 = []
		list22 = []
		list23 = []
		list24 = []
		list25 = []
		list26 = []
		list27 = []
		list28 = []
		list29 = []
		list30 = []
		list31 = []
		list32 = []
		i=1

		for j1 in channelList:
			# if list2[i1] == '2' and (list5[i1] == str(j2)):
			# list0.append(re.findall("<Prj>(.*?)</Prj>", j1, re.S)[0])  # 把这列的数赋值给list1[]
			# list1.append(re.findall("<Yrs>(.*?)</Yrs>", j1, re.S)[0])
			# list2.append(re.findall("<Hyd>(.*?)</Hyd>", j1, re.S)[0])
			# list3.append(re.findall("<sID>(.*?)</sID>", j1, re.S)[0])
			# list4.append(re.findall("<dID>(.*?)</dID>", j1, re.S)[0])
			# list5.append(re.findall("<Flg>(.*?)</Flg>", j1, re.S)[0])
			# list6.append(re.findall("<H1>(.*?)</H1>", j1, re.S)[0])
			# list7.append(re.findall("<H2>(.*?)</H2>", j1, re.S)[0])
			# list8.append(re.findall("<H3>(.*?)</H3>", j1, re.S)[0])
			# list9.append(re.findall("<H4>(.*?)</H4>", j1, re.S)[0])
			# list10.append(re.findall("<H5>(.*?)</H5>", j1, re.S)[0])
			# list11.append(re.findall("<H6>(.*?)</H6>", j1, re.S)[0])
			# list12.append(re.findall("<H7>(.*?)</H7>", j1, re.S)[0])
			# list13.append(re.findall("<H8>(.*?)</H8>", j1, re.S)[0])
			# list14.append(re.findall("<H9>(.*?)</H9>", j1, re.S)[0])
			# list15.append(re.findall("<H10>(.*?)</H10>", j1, re.S)[0])
			# list16.append(re.findall("<H11>(.*?)</H11>", j1, re.S)[0])
			# list17.append(re.findall("<H12>(.*?)</H12>", j1, re.S)[0])
			# list18.append(re.findall("<H13>(.*?)</H13>", j1, re.S)[0])
			# list19.append(re.findall("<H14>(.*?)</H14>", j1, re.S)[0])
			# list20.append(re.findall("<H15>(.*?)</H15>", j1, re.S)[0])
			# list21.append(re.findall("<H16>(.*?)</H16>", j1, re.S)[0])
			# list22.append(re.findall("<H17>(.*?)</H17>", j1, re.S)[0])
			# list23.append(re.findall("<H18>(.*?)</H18>", j1, re.S)[0])
			# list24.append(re.findall("<H19>(.*?)</H19>", j1, re.S)[0])
			# list25.append(re.findall("<H20>(.*?)</H20>", j1, re.S)[0])
			# list26.append(re.findall("<H21>(.*?)</H21>", j1, re.S)[0])
			# list27.append(re.findall("<H22>(.*?)</H22>", j1, re.S)[0])
			# list28.append(re.findall("<H23>(.*?)</H23>", j1, re.S)[0])
			# list29.append(re.findall("<H24>(.*?)</H24>", j1, re.S)[0])
			# list30.append(re.findall("<RR>(.*?)</RR>", j1, re.S)[0])
			# list31.append(re.findall("<SR>(.*?)</SR>", j1, re.S)[0])

			# start = time.clock()
			list0=re.findall("<Prj>(.*?)</Prj>", j1, re.S)[0]  # 把这列的数赋值给list1[]
			list1=re.findall("<Yrs>(.*?)</Yrs>", j1, re.S)[0]
			list2=re.findall("<Hyd>(.*?)</Hyd>", j1, re.S)[0]
			list3=re.findall("<sID>(.*?)</sID>", j1, re.S)[0]
			list4=re.findall("<dID>(.*?)</dID>", j1, re.S)[0]
			list5=re.findall("<Flg>(.*?)</Flg>", j1, re.S)[0]
			list6=re.findall("<H1>(.*?)</H1>", j1, re.S)[0]
			list7=re.findall("<H2>(.*?)</H2>", j1, re.S)[0]
			list8=re.findall("<H3>(.*?)</H3>", j1, re.S)[0]
			list9=re.findall("<H4>(.*?)</H4>", j1, re.S)[0]
			list10=re.findall("<H5>(.*?)</H5>", j1, re.S)[0]
			list11=re.findall("<H6>(.*?)</H6>", j1, re.S)[0]
			list12=re.findall("<H7>(.*?)</H7>", j1, re.S)[0]
			list13=re.findall("<H8>(.*?)</H8>", j1, re.S)[0]
			list14=re.findall("<H9>(.*?)</H9>", j1, re.S)[0]
			list15=re.findall("<H10>(.*?)</H10>", j1, re.S)[0]
			list16=re.findall("<H11>(.*?)</H11>", j1, re.S)[0]
			list17=re.findall("<H12>(.*?)</H12>", j1, re.S)[0]
			list18=re.findall("<H13>(.*?)</H13>", j1, re.S)[0]
			list19=re.findall("<H14>(.*?)</H14>", j1, re.S)[0]
			list20=re.findall("<H15>(.*?)</H15>", j1, re.S)[0]
			list21=re.findall("<H16>(.*?)</H16>", j1, re.S)[0]
			list22=re.findall("<H17>(.*?)</H17>", j1, re.S)[0]
			list23=re.findall("<H18>(.*?)</H18>", j1, re.S)[0]
			list24=re.findall("<H19>(.*?)</H19>", j1, re.S)[0]
			list25=re.findall("<H20>(.*?)</H20>", j1, re.S)[0]
			list26=re.findall("<H21>(.*?)</H21>", j1, re.S)[0]
			list27=re.findall("<H22>(.*?)</H22>", j1, re.S)[0]
			list28=re.findall("<H23>(.*?)</H23>", j1, re.S)[0]
			list29=re.findall("<H24>(.*?)</H24>", j1, re.S)[0]
			list30 = re.findall("<RR>(.*?)</RR>", j1, re.S)[0]
			list31 = re.findall("<SR>(.*?)</SR>", j1, re.S)[0]
			# for i in range(1,len(list29)):
			# 	list31[i]=float(list30[i])+float(list31[i])
			list32 = float(list30) + float(list31)
			if list2 == '2' and 1<=float(list5)<=12 and rank_Hj()[0]<=float(list3)<=rank_Hj()[1]:
				i=i+1
				sheet1.cell(i, 1, self.s)
				sheet1.cell(i, 2, list5)
				sheet1.cell(i, 3, list6)
				sheet1.cell(i, 4, list7)
				sheet1.cell(i, 5, list8)
				sheet1.cell(i, 6, list9)
				sheet1.cell(i, 7, list10)
				sheet1.cell(i, 8, list11)
				sheet1.cell(i, 9, list12)
				sheet1.cell(i, 10, list13)
				sheet1.cell(i, 11, list14)
				sheet1.cell(i, 12, list15)
				sheet1.cell(i, 13, list16)
				sheet1.cell(i, 14, list17)
				sheet1.cell(i, 15, list18)
				sheet1.cell(i, 16, list19)
				sheet1.cell(i, 17, list20)
				sheet1.cell(i, 18, list21)
				sheet1.cell(i, 19, list22)
				sheet1.cell(i, 20, list23)
				sheet1.cell(i, 21, list24)
				sheet1.cell(i, 22, list25)
				sheet1.cell(i, 23, list26)
				sheet1.cell(i, 24, list27)
				sheet1.cell(i, 25, list28)
				sheet1.cell(i, 26, list29)
				sheet1.cell(i, 27, link(list3))
				sheet1.cell(i, 28, list32)
				sheet1.cell(i, 29, list4)
		print(1)
		# for i in range():
		# 	for j in range():
		# 		if sheet1.cell(i,j).value*sheet.cell(i,j).value:

		# 修改excel表单元格格式
		# col0 = sheet1.col(0)
		# col1 = sheet1.col(1)
		# row1 = sheet1.row(0)
		# col0.width = 256 * 30
		# col1.width = 256 * 20
		# row1.set_style(tall_style)
		# book.save('C:\\Users\zhang\Desktop\软件\调试用1126\中间数据\\场景曲线.xls')

		# yingkui=[]  #各小时电力盈亏
		# for s in range(5):#5个分区
		# 	for d in range(1, 4 + 9):  # NS+8
		# 		for j in range(1,13):    # 循环次数12个月
		# 			for i1 in range(len(list1)):#一行的值
		# 				# if list3[i1] == '0' and (down <= float(list4[i1]) <= up) and (list5[i1] == '20'+str(j2) or list5[i1] =='2'+str(j2)):
		# 				if list2[i1] == '2' and (list5[i1] == str(j)):
		# 					yingkui.append(float(self.s))#系统ID
		# 					yingkui.append(float(list5[i1]))#月份
		# 					yingkui.append(float(list6[i1]))#1-24小时
		# 					yingkui.append(float(list7[i1]))
		# 					yingkui.append(float(list8[i1]))
		# 					yingkui.append(float(list9[i1]))
		# 					yingkui.append(float(list10[i1]))
		# 					yingkui.append(float(list11[i1]))
		# 					yingkui.append(float(list12[i1]))
		# 					yingkui.append(float(list13[i1]))
		# 					yingkui.append(float(list14[i1]))
		# 					yingkui.append(float(list15[i1]))
		# 					yingkui.append(float(list16[i1]))
		# 					yingkui.append(float(list17[i1]))
		# 					yingkui.append(float(list18[i1]))
		# 					yingkui.append(float(list19[i1]))
		# 					yingkui.append(float(list20[i1]))
		# 					yingkui.append(float(list21[i1]))
		# 					yingkui.append(float(list22[i1]))
		# 					yingkui.append(float(list23[i1]))
		# 					yingkui.append(float(list24[i1]))
		# 					yingkui.append(float(list25[i1]))
		# 					yingkui.append(float(list26[i1]))
		# 					yingkui.append(float(list27[i1]))
		# 					yingkui.append(float(list28[i1]))
		# 					yingkui.append(float(list29[i1]))
		# 					yingkui.append(float(list3[i1]))
		# 					yingkui.append(float(list30[i1])+float(list31[i1]))
		# 					yingkui.append(float(list4[i1]))
		# 				for h in range(len(yingkui)):
		# 					sheet1.cell(s*(4+8)*12+(d-1)*12+j,h+1,yingkui[h])
		# 			yingkui=[]
		book.save(Path.filepath+'中间数据\\'+'Hust曲线.xlsx')
		# end = time.clock()
        #
		# 		yingkui = []  # 各小时弃水调峰电力置空集，以便下个月的循环
        #
        #
		# 	title1 = ['项目', '1月', '2月', '3月', '4月', '5月', '6月', '7月', '8月', '9月', '10月', '11月', '12月', '全年']
        #
		# 	y = book1.sheet_by_index(ii)
		# 	x = book2.get_sheet(ii)
		# 	for m in range(1, 14):
		# 		ZQDDL.append(y.cell(6, m).value)
		# 	print(ZQDDL)
		# TFBZDL = []  # 调峰不足平均电力期望值指标
		# for k in range(len(TFBZXS)):
		# 	if float(TFBZXS[k]) == 0:
		# 		TFBZDL.append(0)
		# 	else:
		# 		TFBZDL.append(ZQDDL[k] / TFBZXS[k])
        #
		# for i5 in range(len(TFBZDL)):
		# 	cun0.append(TFBZDL[i5])
		# # print(cun0)
        #
		# for i in range(len(title1)):
		# 	sheet1.write(2, i, title1[i])
		# for a in range(3, 3 + mm):
		# 	sheet1.write(a, 0, title2[a - 3])
		# # 数取完了之后，填表
		# i4 = 0
		# for i2 in range(3, 3 + mm):  # 3 4 5 6    #用了编号即方案数！！！
		# 	for i3 in range(1, 14):  # 1-13随机数TIAOFENGBZHuiZ()
		# 		sheet1.write(i2, i3, cun0[i4])  # 把新增写到excel表中
		# 		i4 = i4 + 1
        #
		# self.book.save(Path.filepath1 + jzb1[0] + '系统' + nian + '年' + '调峰不足平均电力期望值汇总表' + '.xls')
        #
		# adapt = xlrd.open_workbook(Path.filepath1 + jzb1[0] + '系统' + nian + '年' + '调峰不足平均电力期望值汇总表' + '.xls','w+')  # 以读的方式打开“dlbz11” 算全年
		# a_adapt = adapt.sheet_by_index(0)  # 读
		# b_adapt = copy(adapt)  # 可写
		# c_adapt = b_adapt.get_sheet(0)  # 写
		# c_adapt.write_merge(0, 0, 0, 13, '表' + str(1) + '    ' + jzb1[0] + '系统' + nian + '调峰不足平均电力期望值汇总表')
		# c_adapt.write_merge(1, 1, 0, 13, '( ' + beizhu + "     单位："+"GW"+')')
		# b_adapt.save(Path.filepath1 + jzb1[0] + '系统' + nian + '年' + '调峰不足平均电力期望值汇总表' + '.x



# s=0
# exchange(s,wpl)