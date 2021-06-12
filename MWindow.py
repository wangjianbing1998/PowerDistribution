import os
import sys

import numpy as np
import openpyxl
import pandas as pd
from PyQt5.QtWidgets import *

# import run
import Path
import Sysmat
import check
import solver
from MWin import Ui_MainWindow
from Opr_setting import SWindow


# import clr  # clr是公共运行时环境，这个模块是与C#交互的核心
# 将self.orderDict中的信息写入本地xml文件，参数filename是xml文件名
# clr.FindAssembly("lei.dll")  ## 加载c#dll文件
# from lei import * # 导入命名空间
# clr.FindAssembly("HUST_CMD.dll")  ## 加载c#dll文件
# from HUST_CMD import *   # 导入命名空间

# 这一版是把每张excel表转化成同等大小的矩阵形式，如果有矩阵计算就比较方便
class YWindow(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.setWindowTitle("主窗口")
        self.action1.triggered.connect(self.open)
        self.action3_2.triggered.connect(self.set)
        self.action2.triggered.connect(self.exit)
        # self.action3.triggered.connect(self.curve)
        self.action4.triggered.connect(self.DC)
        # self.action4.triggered.connect(self.AC)

    def open(self):
        file_name, ok = QFileDialog.getOpenFileName(self, "打开", "D:/", "Excel File(*.xlsx)")
        if file_name == '':
            return
        Path.filepath = os.path.dirname(file_name) + '/'
        Path.filename = os.path.basename(file_name)
        #        self.lineEdit.setText(Path.filepath + Path.filename)
        print(Path.filepath)
        # 以下检查文件夹是否齐全
        dirct = Path.filepath  # 总文件夹路径
        dirList = []  # 存储总文件夹下的所有目录名
        fileList = []  # 存储总文件夹下的所有文件名
        files = os.listdir(dirct)  # 文件夹下所有目录的列表
        for f in files:
            if os.path.isdir(dirct + '/' + f):  # 这里是绝对路径，该句判断目录是否是文件夹
                dirList.append(f)
            elif os.path.isfile(dirct + '/' + f):  # 这里是绝对路径，该句判断目录是否是文件
                fileList.append(f)
        n = ['Hust数据', '日历曲线', '输出数据', '中间数据']
        if check.findNum(dirList, n) != [] or check.findNum(n, dirList) != []:
            QMessageBox.information(self, "错误", "目录缺失或不正确", QMessageBox.Yes)
            tip1 = "打开失败"
            self.statusbar.showMessage(tip1)
        else:
            # 以下检查南疆电网-子表的名称和顺序是否正确
            Path.wh = openpyxl.load_workbook(Path.filepath + 'Hust数据' + '/南疆电网.xlsx')
            wh=Path.wh
            sheets0 = wh.sheetnames
            truename=['线路表','机组表','日曲线','年曲线','系统表','方案表']
            print(sheets0)
            if sheets0!=truename:
                QMessageBox.information(self, "错误", "’南疆电网‘子表名称或顺序不正确！\n正确顺序应为: 线路表-机组表-日曲线-年曲线-系统表-方案表", QMessageBox.Yes)
                tip1 = "打开失败"
                self.statusbar.showMessage(tip1)
            else:
                # 以下检查算例-子表是否齐全（不多不少）
                wb = openpyxl.load_workbook(Path.filepath + Path.filename)
                # wb=openpyxl.load_workbook('C:\\Users\zhang\Desktop\软件\调试用\中间数据\场景曲线.xlsx')
                sheets = wb.sheetnames
                m = ['基础表', '系统表', '外送特性', '线路表', '电站表', '特性表']
                if check.findNum(sheets, m) != [] or check.findNum(m, sheets) != []:  #
                    QMessageBox.information(self, "错误", "主文件子表缺失或不正确", QMessageBox.Yes)
                    # if
                else:
                    for name in m:
                        table = wb.get_sheet_by_name(name)  ##根据名称获取某个sheet
                        rows = table.max_row  ##3获取当前sheet页面的总行数,把每一行数据作为list放到 list
                        cols = table.max_column  ##获取当前sheet页面的总列数
                        result = []  # Excel数据存到二维列表，包括表头
                        col = []
                        for j in range(1, cols + 1):
                            col.append(table.cell(1, j).value)
                        result.append(col)  # 单独读表头
                        for i in range(2, rows + 2):
                            col = []
                            for j in range(1, cols + 1):
                                colvalue = table.cell(i, j).value
                                col.append(colvalue)
                            if type(col[0]) == int and type(col[1]) == str:  # 只选取有效行
                                result.append(col)
                        list_to_matrix = np.mat(result)  # 列表转矩阵,转成与excel表格大小相同的矩阵
                        if name == '基础表':
                            Path.base = list_to_matrix
                        elif name == '系统表':
                            Path.system = list_to_matrix
                        elif name == '线路表':
                            Path.line = list_to_matrix
                            Path.linetable = table
                        elif name == '电站表':
                            Path.station = list_to_matrix
                        elif name == '特性表':
                            Path.character = list_to_matrix
                        elif name == '外送特性':
                            Path.waisong = list_to_matrix
                G = Sysmat.getmatrix(Path.linetable).G  # 调用生成G逆矩阵的模块并存储为全局变量
                try:
                    Path.matrix = G.I
                    tip1 = "打开成功"
                except np.linalg.LinAlgError:
                    QMessageBox.information(self, "错误", "节点电导矩阵奇异，系统存在孤立节点。", QMessageBox.Yes)
                    tip1 = "打开失败"
                self.statusbar.showMessage(tip1)
    def set(self):#读Hust安装位置
        self.Setting_Item=SWindow()
        self.Setting_Item.Visi()

    def DC(self):
        # 新能源数据准备
        wh = Path.wh
        sheet_plant = pd.read_excel(
            os.path.join(os.path.dirname(__file__), Path.filepath + Path.filename), sheet_name='电站表'
        )
        sheet_hustline = pd.read_excel(
            os.path.join(os.path.dirname(__file__), Path.filepath + 'Hust数据' + '/南疆电网.xlsx'), sheet_name='线路表'
        )
        wb = openpyxl.load_workbook(Path.filepath + '中间数据' + '/场景曲线.xlsx')
        wst = wb.get_sheet_by_name('电站出力')  # 调电站出力表，待下面写入数据
        wpl = wb.get_sheet_by_name('电力盈亏')  # 调电力盈亏表，待下面写入数据
        wex = wb.get_sheet_by_name('优化交换')  # 调优化交换表，待下面写入数据
        whust = wh.get_sheet_by_name('日曲线')  # 调南疆数据的日曲线表，待下面写入数据
        ss = []  # 节点数
        for i in range(Path.system.shape[0]):  # 矩阵行数
            if Path.system[i, 2] == str(100) or Path.system[i, 2] == str(101):
                ss.append(int(Path.system[i, 0]))
        ll = []  # 联络线ID
        l_hust = []  # 联络线在Hust中的ID
        for i in range(Path.line.shape[0]):
            if Path.line[i, 6] == str(400):
                ll.append(int(Path.line[i, 0]))
                l_hust.append(int(Path.line[i, 5]))
        kk = []  # 火电站数
        for i in range(Path.station.shape[0]):
            if Path.station[i, 5] == 101:
                kk.append(int(Path.station[i, 0]))
        singlecapacity = []  # 读各电站单机容量
        techPower = []  # 技术出力
        totalNums = []  # 台数
        for _, plant_row_values in sheet_plant.iterrows():
            if plant_row_values["电站类型"] == 101:
                singlecapacity.append(plant_row_values["单机容量"])  # 火电单机
                techPower.append(plant_row_values["技术出力"])
                totalNums.append(plant_row_values["台数"])
        PTmin = np.array(singlecapacity) * np.array(techPower)  # 最小技术出力
        TotalCapacity = np.array(singlecapacity) * np.array(totalNums)  # 装机

        ConnectlineCapacity = []  # 联络线容量
        ConnectlineCapacityID = []
        for _, hustline_row_values in sheet_hustline.iterrows():
            if hustline_row_values["回路数"] == 1:
                ConnectlineCapacity.append(hustline_row_values["单回Pmax"])
                ConnectlineCapacityID.append(hustline_row_values["线路ID"])  # 联络线ID
        # *****************主程序***********************#
        # 调用上层直流计算程序
        DC_operation = solver.Solver()
        DC_operation.solve()
        result = DC_operation.afterProcess()
        for s in range(0, 5):  # s场景：0,1,2,3,4 后续改成用户输入的场景编号
            for m in range(1, 13):  # m月份：1,2,……,12 后续2改成13
                # **************************一个单位日所在循环*******************************#
                # (UCT,PT,inject,abandon)=UC.Unit_Commitment(Path.system,Path.station,Path.character,Path.waisong,PV_joint,PV_station,WS)#后续如需增减参数可自行设定
                # 分别返回各火电站开机矩阵，UCT：1*12、各火电站出力矩阵，PT：12*24、各节点注入功率矩阵，inject：6*24
                # (UCnew, PTnew, yingkui)=check.PLine(Path.g,Path.matrix,Path.system,Path.line,Path.station,UCT,PT,inject)
                # 以下for循环将火电站出力写入中间数据表的'电站出力'sheet中
                Result = result[s, m]
                openSchema = np.array(Result.openSchemaOnStation)[:, 1]  # 开机
                PT = np.array(Result.thermalPowerOnStation)  # 火电出力
                yingkui = np.array(Result.abandonOnNode)

                KJ_capacity = openSchema * PTmin  # 开机容量
                Reserve = 0.1 * np.array(singlecapacity)  # 旋转备用
                Repair = np.zeros((12,))  # 检修容量
                PT = np.column_stack(
                    (PT, TotalCapacity, KJ_capacity, PTmin, Reserve, Repair))
                for i in range(len(kk)):  # 火电站的数量
                    wst.cell(s * 144 + i * 12 + m + 1, 1, s)  # 场景编号写入电站出力表
                    wst.cell(s * 144 + i * 12 + m + 1, 2, m)  # 月份写入
                    for j in range(30):
                        wst.cell(s * 144 + i * 12 + m + 1, j + 3, PT[i, j])  # 出力+电站ID写入
                #以下for循环将电力盈亏写入中间数据表的'电力盈亏'sheet中
                for i in range(len(ss)):#节点/系统数
                    wpl.cell(s * 72 + i * 12 + m + 1, 1, s)  # 场景编号写入电站出力表
                    wpl.cell(s * 72 + i * 12 + m + 1, 2, m)  # 月份写入
                    for j in range(25):
                        wpl.cell(s * 72 + i * 12 + m + 1, j + 3, yingkui[i, j])  # 盈亏+系统ID写入
                #以下for循环将交换功率写入中间数据表的'优化交换'sheet中
                for i in range(len(ll)):#线路数
                    wex.cell(s * 72 + i * 12 + m + 1, 1, s)  # 场景编号写入电站出力表
                    wex.cell(s * 72 + i * 12 + m + 1, 2, m)  # 月份写入
                    wex.cell(s * 72 + i * 12 + m + 1, 27, ll[i])  # 线路ID写入
                    wex.cell(s * 72 + i * 12 + m + 1, 29, l_hust[i])  # HustID写入
                    for j in range(24):
                        wex.cell(s * 72 + i * 12 + m + 1, j + 3, yingkui[i, j])  # 盈亏写入
            #以下for循环将联络线功率写入南疆电网的'日曲线'sheet中，该循环处于月循环之外
            flag = 0
            kend = 0
            for k in range(2, 230):  # 230待一般化
                if flag == 0:
                    for t in range(s * (6 * 12) + 2, (s + 1) * (6 * 12) + 2):
                        if int(wex.cell(t, 29).value) == int(whust.cell(k, 2).value):
                            flag = 1
                            for id in ConnectlineCapacityID:
                                if int(wex.cell(t, 29).value) == id:
                                    lineid = ConnectlineCapacityID.index(id)
                            print(lineid)
                            for m in range(12):
                                for j in range(24):
                                    whust.cell(k + m, j + 9,
                                               min(wex.cell(t + m, j + 3).value, ConnectlineCapacity[lineid]) /
                                               ConnectlineCapacity[lineid])  # 联络线数据写入
                                kend = k + m
                            break
                if k == kend:
                    flag = 0
            #********excel转xml********#
            #xmlfile=EtoX.toxml(wh,whust)注意：xml文件的名字
            #xmlfile=ETOX11.writeInfoToXml11()
            #尽量不使用保存（耗时），直接调用中间变量。wh是整个excel，后续可在该函数中打开需要的sheet；whust是上一步更新过的日曲线sheet，单独作为一个参数传入
            # #调用Hust_Pros
            # run.run_Hust(xmlfile)#调用Hust_Pros
            #上下层功率交换
            # Hustoutput1223.exchange(s,wpl,wex)
            print(111111)

        wb.save(Path.filepath + '中间数据' + '/场景曲线.xlsx')
        wh.save(Path.filepath + 'Hust数据' + '/南疆电网.xlsx')


    def exit(self):
        self.close()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    myWin = YWindow()
    myWin.show()
    sys.exit(app.exec())

