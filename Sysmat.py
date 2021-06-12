import numpy as np
import Path
import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\zhang\Desktop\软件\调试用\算例.xlsx')
line=wb.get_sheet_by_name('线路表')
system=wb.get_sheet_by_name('系统表')
rows = system.max_row  ##3获取当前sheet页面的总行数,把每一行数据作为list放到 list
cols = system.max_column  ##获取当前sheet页面的总列数
result = []  # Excel数据存到二维列表，包括表头
col = []
for j in range(1, cols + 1):
    col.append(system.cell(1, j).value)
result.append(col)  # 单独读表头
for i in range(2, rows + 2):
    col = []
    for j in range(1, cols + 1):
        colvalue = system.cell(i, j).value
        col.append(colvalue)
    if type(col[0]) == int and type(col[1]) == str:  # 只选取有效行
        result.append(col)
list_to_matrix = np.mat(result)  # 列表转矩阵,转成与excel表格大小相同的矩阵
Path.system=list_to_matrix
class getmatrix:
    def __init__(self,A):
        self.A=A
        def search_G(S,S_D):  # 查找与节点S直接相连的节点电导
            M = []#相连的节点
            ID = []#对应的线路ID
            for i in range(1,len(S_D)+2):
                if self.A.cell(i,self.SDlocation).value ==S:
                    M.append(self.A.cell(i,self.MDlocation).value)
                    ID.append(self.A.cell(i,self.lIDlocation).value)
            for i in range(1,len(S_D)+2):
                if self.A.cell(i,self.MDlocation).value ==S:
                    M.append(self.A.cell(i,self.SDlocation).value)
                    ID.append(self.A.cell(i, self.lIDlocation).value)
            return (M,ID)

        def calculate_Gline(line_ID):#根据线路ID求线路电导
            for i in range(self.rows-1):
                print(i)
                if self.A.cell(i+2,self.lIDlocation).value==line_ID:
                    if self.A.cell(i+2,self.Numlocation).value%2==0:#回路数是2的倍数
                        r=self.A.cell(i+2,self.lengthlocation).value*self.A.cell(i+2,self.Rlocation).value*2
                        print(r)
                        if r==0:
                            g=pow(10,18)
                        else:
                            g=(1/r)*(self.A.cell(i+2,self.Numlocation).value/2)
                        print(g)
                    if self.A.cell(i+2,self.Numlocation).value%2==1:#回路数不是2的倍数
                        r=self.A.cell(i+2,self.lengthlocation).value*self.A.cell(i+2,self.Rlocation).value
                        if r==0:
                            g=pow(10,18)
                        else:
                            g=1/r+(1/(r*2))*(self.A.cell(i+2,self.Numlocation).value//2)
            return g

        # 加一个判断，防止读到空行和空列
        self.rows=0
        for i in range(1,self.A.max_row+1):
            if self.A.cell(i,1).value != 'None':
                self.rows=self.rows+1
        self.cols=0
        for i in range(1,self.A.max_column+1):
            if self.A.cell(2,i).value != 'None':
                self.cols=self.cols+1
        N = []
        L = []
        R = []
        S_D = []#首端节点
        M_D = []#末端节点
        upID = []#存储属于上层直流的线路ID
        for i in range(self.cols):
            if self.A.cell(1,i+1).value=='类型':
                for j in range(self.rows-1):
                    if self.A.cell(j+2, i + 1).value==410:
                        upID.append(self.A.cell(j+2,1).value)
        for i in range(self.cols):
            if self.A.cell(1,i+1).value=='回路数':
                self.Numlocation = i + 1  # 记录线路回路数所在的列
                for j in range(self.rows-1):
                    if self.A.cell(j+2, 1).value in upID:
                        N.append(self.A.cell(j+2,i+1).value)
            elif self.A.cell(1,i+1).value=='线路ID':
                self.lIDlocation=i+1#记录线路ID所在的列
            elif self.A.cell(1,i+1).value=='长度':
                self.lengthlocation = i+1#记录线路长度所在的列
                for j in range(self.rows-1):
                    if self.A.cell(j + 2, 1).value in upID:
                        L.append(self.A.cell(j+2,i+1).value)
            elif self.A.cell(1,i+1).value=='电阻':
                self.Rlocation = i + 1  # 记录线路电阻所在的列
                for j in range(self.rows-1):
                    if self.A.cell(j + 2, 1).value in upID:
                        R.append(self.A.cell(j+2,i+1).value)
            elif self.A.cell(1,i+1).value=='首端ID':
                self.SDlocation=i+1#记录首端ID是所在列
                for j in range(self.rows-1):
                    if self.A.cell(j + 2, 1).value in upID:
                        S_D.append(self.A.cell(j+2,i+1).value)
            elif self.A.cell(1,i+1).value=='末端ID':
                self.MDlocation = i + 1  # 记录末端ID所在列
                for j in range(self.rows-1):
                    if self.A.cell(j + 2, 1).value in upID:
                        M_D.append(self.A.cell(j+2,i+1).value)
    
        for j in range(Path.system.shape[1]):
            if Path.system[0, j] == '系统ID':
                self.SysID_location=j
    
        self.Sys_ID = []#存储上层所有系统ID
        self.G_ID = []  # 存储上层非松弛节点（0节点）系统ID
        print(1)
        for j in range(Path.system.shape[1]):
            if Path.system[0,j]=='类型':
                for i in range(1,Path.system.shape[0]):
                    if int(Path.system[i,j])==101 or int(Path.system[i,j])==100:
                        self.Sys_ID.append(int(Path.system[i,self.SysID_location]))
                    if int(Path.system[i, j]) == 101 :  #
                        self.G_ID.append(int(Path.system[i, self.SysID_location]))
        self.G = np.mat(np.zeros((len(self.G_ID), len(self.G_ID))))
        print(self.G)
        # print(type(G_ID))
        for i in self.G_ID:#对每个节点求自电导和互电导
            x =self.G_ID.index(i)
            (node_ID,line_ID)=search_G(i,S_D)#node_ID是与i相连的节点ID，line_ID是对应的线路ID
            for k in line_ID:
                self.G[x, x] = self.G[x, x] + calculate_Gline(k)#自电导
                if self.G[x, x]>pow(10,17):
                    self.G[x, x]=0
            for j in self.G_ID :
                y = self.G_ID.index(j)
                if j!=i:
                    if j not in node_ID:#互电导
                        self.G[x, y]=0
                    else:
                        self.G[x,y]=-calculate_Gline(line_ID[node_ID.index(j)])
                        if abs(self.G[x, y]) > pow(10, 17):
                            self.G[x, y] = 0

        print(1/self.G)

#返回5*5的G逆矩阵

getmatrix(line)